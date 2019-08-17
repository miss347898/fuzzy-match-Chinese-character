from bank import list_bank
from lender import list_lender

from fuzzywuzzy import fuzz
from fuzzywuzzy import process

import openpyxl

list_match = []
list_score = []
for lender in list_lender:
   matched = process.extractOne(lender, list_bank)[0]
   score = process.extractOne(lender, list_bank)[1]
   list_match.append(matched)
   list_score.append(score)

workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = "matched"

for i in range(len(list_match)):
    worksheet.cell(i + 1, 1, list_match[i])
    worksheet.cell(i + 1, 2, list_score[i])
workbook.save(filename=r'C:\Users\347898222\Desktop\编程\excel项目\matched.xlsx')