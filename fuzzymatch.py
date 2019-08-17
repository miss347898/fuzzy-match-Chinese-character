import pandas as pd
import openpyxl
from fuzzywuzzy import fuzz
from fuzzywuzzy import process
import xlwings as xw

#关闭现有的excel控制台
try:
    wb.close()
    app.quit()
except:
    pass



#读入stata文件
def load_large_dta(fname):
    import sys

    reader = pd.read_stata(fname, iterator=True)
    df = pd.DataFrame()

    try:
        chunk = reader.get_chunk(100 * 1000)
        while len(chunk) > 0:
            df = df.append(chunk, ignore_index=True)
            chunk = reader.get_chunk(100 * 1000)
            print('.')
            sys.stdout.flush()
    except (StopIteration, KeyboardInterrupt):
        pass

    print('\nloaded {} rows'.format(len(df)))

    return df

def deconde_str(string):

    re = string.encode('latin-1').decode('utf-8')
    return re

# read
lender_origin_load = r"C:\Users\347898222\Desktop\编程\excel项目/lender_origin2.dta"
lender_origin = load_large_dta(lender_origin_load)

result = r"C:\Users\347898222\Desktop\编程\excel项目"
print(lender_origin['bank_type'][1])


for type_bank in range(0, 8):
    lender_list = []
    for i in range(0, 45970):
        if lender_origin['bank_type'][i] == type_bank:
            lender_name = lender_origin['lender'][i]
            lender_list.append(lender_name)

    app = xw.App(visible=True, add_book=False)
    wb = app.books.add()
    wb.sheets['sheet1'].range('A1').options(transpose=True).value = lender_list

    wb.save(r'C:\Users\347898222\Desktop\编程\excel项目\lender_%s.xlsx' % type_bank)
    wb.close()
    app.quit()


#lender名单提取
for i in range(0,8):
    try:
        workbook = openpyxl.load_workbook(r"C:\Users\347898222\Desktop\编程\excel项目\modify\lender_%s.xlsx" % i)
        shenames = workbook.sheetnames
        worksheet = workbook[shenames[0]]
        list_lender = []
        list_id = []
        list_type = []
        for cell in list(worksheet.columns)[0]:
            nameoflender = cell.value
            list_lender.append(nameoflender)
        print(len(list_lender))
    except:
        pass

#银行名单提取
    try:
        list_bank = []
        workbook = openpyxl.load_workbook(r"C:\Users\347898222\Desktop\编程\excel项目\modify\bank_%s.xlsx" % i)
        shenames = workbook.sheetnames
        worksheet = workbook[shenames[0]]
        for cell in list(worksheet.columns)[0]:  # 获取第三列的数据
            nameofbank = cell.value
            list_bank.append(nameofbank)
        print(len(list_bank))
    except:
        pass

#开始匹配
    list_match = []
    list_score = []
    try:
        for lender in list_lender:
            matched = process.extractOne(lender, list_bank)[0]
            score = process.extractOne(lender, list_bank)[1]
            list_match.append(matched)
            list_score.append(score)
    except:
        pass
    print(i)

#生成最终表格
    app = xw.App(visible=True, add_book=False)
    wb = app.books.add()
    wb.sheets['sheet1'].range('A1').options(transpose=True).value = list_lender
    wb.sheets['sheet1'].range('B1').options(transpose=True).value = list_match
    wb.sheets['sheet1'].range('C1').options(transpose=True).value = list_score

    wb.save(r'C:\Users\347898222\Desktop\编程\excel项目\modify\matched_%s.xlsx' % i)
    wb.close()
    app.quit()
#
#
#
# #lender2单独处理
# workbook = openpyxl.load_workbook(r"C:\Users\347898222\Desktop\编程\excel项目\lender_2_1.xlsx" )
# shenames = workbook.sheetnames
# worksheet = workbook[shenames[0]]
# list_lender = []
# list_id = []
# list_type = []
# for cell in list(worksheet.columns)[0]:
#     nameoflender = str(cell.value)
#     list_lender.append(nameoflender)
# print(len(list_lender))
#
# #银行名单提取
# list_bank = []
# workbook = openpyxl.load_workbook(r"C:\Users\347898222\Desktop\编程\excel项目\bank_2.xlsx" )
# shenames = workbook.sheetnames
# worksheet = workbook[shenames[0]]
# for cell in list(worksheet.columns)[0]:
#     nameofbank = cell.value
#     list_bank.append(nameofbank)
# print(len(list_bank))
#
# #开始匹配
# list_match = []
# list_score = []
# try:
#     for lender in list_lender:
#         matched = process.extractOne(lender, list_bank)[0]
#         score = process.extractOne(lender, list_bank)[1]
#         list_match.append(matched)
#         list_score.append(score)
# except:
#     pass
#
# app = xw.App(visible=True, add_book=False)
# wb = app.books.add()
# wb.sheets['sheet1'].range('A1').options(transpose=True).value = list_lender
# wb.sheets['sheet1'].range('B1').options(transpose=True).value = list_match
# wb.sheets['sheet1'].range('C1').options(transpose=True).value = list_score
#
# wb.save(r'C:\Users\347898222\Desktop\编程\excel项目\matched_2_1.xlsx' )
# wb.close()
# app.quit()
