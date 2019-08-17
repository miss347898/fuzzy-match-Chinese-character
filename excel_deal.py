import xlrd
import xlsxwriter
import openpyxl


for i in range(0, 8):
 workbook = openpyxl.load_workbook(r"C:\Users\347898222\Desktop\编程\excel项目\modify\matched_%s.xlsx" % i)
 worksheet = workbook.active
 print(worksheet)
 rows = worksheet.max_row
 columns = worksheet.max_column
 print(rows, columns)

 for j in range(1, rows + 1):
  worksheet.cell(j, 4, i)

 workbook.save(r"C:\Users\347898222\Desktop\编程\excel项目\modify\matched_%s.xlsx" % i)


target_xls = r"C:\Users\347898222\Desktop\编程\excel项目\modify\final_merge.xlsx"

# 读取数据
data = []
for i in range(0, 8):
 wb = xlrd.open_workbook(r"C:\Users\347898222\Desktop\编程\excel项目\modify\matched_%s.xlsx" % i)
 for sheet in wb.sheets():
  for rownum in range(sheet.nrows):
   data.append(sheet.row_values(rownum))

# 写入数据
workbook = xlsxwriter.Workbook(target_xls)
worksheet = workbook.add_worksheet()
font = workbook.add_format({"font_size":14})
for i in range(len(data)):
 for j in range(len(data[i])):
  worksheet.write(i, j, data[i][j], font)
# 关闭文件流
workbook.close()