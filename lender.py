
import openpyxl

# 获取 工作簿对象
workbook = openpyxl.load_workbook(r"C:\Users\347898222\Desktop\编程\excel项目\stata.xlsx")
shenames = workbook.sheetnames
print(shenames)

worksheet = workbook[shenames[0]]
print(worksheet)

# 经过上述操作，我们已经获得了第一个“表”的“表对象“，接下来可以对表对象进行操作
name = worksheet.title
rows = worksheet.max_row
columns = worksheet.max_column
print(name, rows, columns)


#for row in worksheet.rows:
#    for cell in row:
#        print(cell.value,end=" ")
#    print()

#for col in worksheet.columns:
#    for cell in col:
#        print(cell.value,end=" ")
#    print()
list_lender = []
for i in range(1, rows+1):
    for j in range(1, columns+1):
        list_lender.append(worksheet.cell(row=i, column=j).value)

print(list_lender)
